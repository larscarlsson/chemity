import pandas as pd
from pathlib import Path
from openpyxl import load_workbook



'''
Read xls or xlsx file to find SCIP and CAS numbers.
'''

repo_path = Path(__file__).resolve().parents[2]
folder_path = repo_path / 'source_data/examples_chemity/Underlag'  # Replace with your folder path

example_xls_path = folder_path / 'capacitor_49322_06162014_061620141.xls'
example_xlsx_path = folder_path / '01_part-list_containing_svhc_revision18_20230428-v1.xlsx' # this contains several tabs. find the right


def find_scip_sheet(excel_file):
    # Open the workbook
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    
    # Get all sheet names
    sheet_names = workbook.sheetnames
    
    # Iterate through sheets to find one with "SCIP" heading
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        # Read first row
        for i, row in enumerate(sheet.iter_rows(max_row=3, values_only=True)):
            # Check if "SCIP" is in the first 3 rows (case-insensitive)
            if any("CAS" in str(cell).upper() for cell in row if cell) and any("SCIP" in str(cell).upper() for cell in row if cell):
                # Read the specific sheet into pandas
                df = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=i, engine="openpyxl")
                return df, sheet_name
    
    # If no sheet with SCIP is found
    raise ValueError("No sheet with 'SCIP' in header found")

    
print(find_scip_sheet(example_xlsx_path))